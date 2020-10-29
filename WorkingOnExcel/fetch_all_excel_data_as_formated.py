'''
@Author:Yohannes Assebe(John Assebe)
This will print the Excel files in their format 
'''
import openpyxl
wb=openpyxl.load_workbook('censuspopdata.xlsx')
shet=wb['Population by Census Tract']
##print(shet.max_column)
for i in range(1,shet.max_row+1):
    for k in range(1,shet.max_column+1):
        if shet.cell(row=i,column=k).value==None:
            shet.cell(row=i,column=k).value=" "
        print(shet.cell(row=i,column=k).value,end="\t\t")
    print()
        
