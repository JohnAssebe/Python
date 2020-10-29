'''
@Author:Yohannes Assebe(John Assebe)
This program inverts the spreadsheet data
'''
import openpyxl,subprocess,logging
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s-%(levelname)s-%(message)s')
wb=openpyxl.load_workbook('invert.xlsx')
sheet=wb.active
for i in range(1,sheet.max_row+1):
    for k in range (1,sheet.max_row+1):
        first=sheet.cell(row=k,column=i).value
##        logging.debug('%s%%',first)
        sheet.cell(row=k,column=i).value=sheet.cell(row=i,column=k).value
##        logging.debug('%s%%',sheet.cell(row=k,column=i).value)
        sheet.cell(row=i,column=k).value=first
##        logging.debug('%s%%',sheet.cell(row=i,column=k).value)
wb.save('inverted.xlsx')
print("Done thanks for using this system")
subprocess.Popen(['start','inverted.xlsx'],shell=True)
subprocess.Popen(['start','invert.xlsx'],shell=True)
