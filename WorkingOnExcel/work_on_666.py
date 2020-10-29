import openpyxl
from openpyxl.styles import Font
import subprocess
wb=openpyxl.load_workbook("table.xlsx",data_only=True)
sheet=wb['Sheet']
fontobj1=Font(name='Calibri',size=19,bold=True)
print(sheet.cell(row=3,column=3).value)
for rows in (1,sheet.max_row+1):
    for columns in range(1,sheet.max_column+1):
        print(sheet.cell(row=rows,column=columns).value)
    print()
##           if sheet.cell(row=a,column=b).value==666:
##                 sheet.cell(row=a,column=b).font=fontobj1
print("done")
##wb.save("john.xlsx")
##subprocess.Popen(['start','john.xlsx'],shell=True)
